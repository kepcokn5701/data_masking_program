"""
N2SF 정보 분류·마스킹 엔진 (공용 코어)
────────────────────────────────────────────────────────────
데스크톱 GUI(excel_masking.py)와 웹서버(web/app.py)가 공유하는 순수 로직.
UI/프레임워크 의존성 없음 — pandas, re 만 사용.

업로드된 엑셀의 모든 셀을 N2SF 기준(C/S/O)으로 분류하고 등급별로 차등 마스킹한다.
  · C(기밀)  : 완전 마스킹   (주민번호·카드·계좌·여권·면허)
  · S(민감)  : 부분 마스킹   (이름·전화·이메일·주소·생년월일·차량)
  · O(공개)  : 원본 유지     (그 외 전부)
한 셀에 여러 정보가 섞여 있어도 각각 탐지·마스킹한다(셀 단위 전체 스캔).

[참고] N2SF는 향후 전환 대비 참고 분류이며, 이 도구의 핵심은 '마스킹'이다.
"""

import io
import re
from datetime import datetime, date, time
import openpyxl

# ── N2SF 근거 (외부 링크 없음 · 원문은 국정원 홈페이지에서 확인하도록 안내) ──
STANDARD_NAME = "국가 망 보안체계(N2SF) 보안가이드라인 1.0 · 국가정보원(NIS)"
STANDARD_TIMELINE = "초안(Draft) 2025-01 → 정식판 1.0 2025-09 (CSK 2025)"
GUIDE_NOTE = "원문은 국정원(국가사이버안보센터) 홈페이지에서 확인할 수 있습니다."
TREND_NOTES = [   # 정식 공문 아님 — 참고 동향정보(외부 링크 미첨부)
    "국정원 보도자료 — N2SF 보안가이드라인 정식판 공개 (2025-09)",
    "ZDNet — 보안 통제 항목 약 170→260개 확대 보도 (2025-09-09)",
    "한국일보 — N2SF 가이드라인 정식판 공개 보도 (2025-09-30)",
]

# ── N2SF 등급 정의 ────────────────────────────────────────────
GRADE_LABEL  = {"C": "기밀(C)", "S": "민감(S)", "O": "공개(O)"}
GRADE_COLOR  = {"C": "#dc2626", "S": "#f59e0b", "O": "#94a3b8"}
GRADE_RANK   = {"C": 0, "S": 1, "O": 2}            # 충돌 해소 우선순위
GRADE_POLICY = {"C": "완전 마스킹", "S": "부분 마스킹", "O": "원본 유지"}
GRADE_DEF = {
    "C": "정보공개법 등 비공개정보 중 중요도가 높은 고유식별·금융정보",
    "S": "공개 시 개인·국가 이익 침해가 우려되는 식별 가능 정보",
    "O": "기밀·민감을 제외한 모든 정보(원본 유지)",
}

# ── 마스킹 함수들 (매칭된 문자열 → 마스킹 문자열) ──────────────

def _mask_all_digits(seg):
    """숫자만 전부 * (구분자·기호는 유지) → 주민/카드/계좌/면허"""
    return re.sub(r"\d", "*", seg)

def _mask_passport(seg):
    """여권번호: 앞 1글자만 남기고 마스킹 (M12345678 → M********)"""
    return seg[0] + "*" * (len(seg) - 1) if seg else seg

def _mask_phone(seg):
    """전화번호: 가운데 국번을 마스킹 (010-****-5678, 02-***-4567)"""
    parts = re.split(r"([-\s]+)", seg)
    groups = [i for i, p in enumerate(parts) if p.isdigit()]
    if len(groups) >= 3:                       # 구분자로 3그룹 이상 → 가운데 마스킹
        for i in groups[1:-1]:
            parts[i] = "*" * len(parts[i])
        return "".join(parts)
    d = re.sub(r"\D", "", seg)                  # 구분자 없는 경우 자릿수로 처리
    if len(d) == 11:
        return f"{d[:3]}-****-{d[7:]}"
    if len(d) == 10:
        return f"{d[:3]}-***-{d[6:]}"
    return re.sub(r"\d(?=\d{4})", "*", seg)     # 그 외: 마지막 4자리만 남김

def _mask_email(seg):
    """이메일: ho**@gmail.com"""
    local, _, domain = seg.partition("@")
    keep = max(2, len(local) // 3)
    return local[:keep] + "*" * (len(local) - keep) + "@" + domain

def _mask_birth(seg):
    """생년월일: 연도만 남기고 월·일 마스킹 (1990-01-01 → 1990-**-**)"""
    m = re.match(r"(\d{4})([.\-/])(\d{1,2})([.\-/])(\d{1,2})", seg)
    if m:
        return f"{m.group(1)}{m.group(2)}**{m.group(4)}**"
    return re.sub(r"\d", "*", seg)

def _mask_vehicle(seg):
    """차량번호: 숫자만 마스킹 (12가3456 → **가****)"""
    return re.sub(r"\d", "*", seg)

def _mask_name(seg):
    """이름: 홍길동 → 홍*동, 김철 → 김*"""
    s = seg.strip()
    if len(s) == 2:
        return s[0] + "*"
    if len(s) >= 3:
        return s[0] + "*" * (len(s) - 2) + s[-1]
    return s

_ADDR_REGION = (r"(?:서울|부산|대구|인천|광주|대전|울산|세종|경기|강원특별자치도|강원|"
                r"충청북도|충청남도|충북|충남|전라북도|전북특별자치도|전라남도|전북|전남|"
                r"경상북도|경상남도|경북|경남|제주특별자치도|제주)"
                r"(?:특별자치시|특별자치도|특별시|광역시|시|도)?")

def _mask_address(seg):
    """주소: 시/도만 남기고 이후 토큰 마스킹 (서울시 강남구 역삼동 → 서울시 *** ***)"""
    m = re.match(_ADDR_REGION, seg)
    region = m.group() if m else ""
    rest = seg[len(region):]
    masked = re.sub(r"[가-힣A-Za-z0-9]+", lambda x: "*" * len(x.group()), rest)
    return region + masked

# ── 형식 검증 함수 (오탐 방지: 일반 숫자코드를 PII로 오인하지 않도록) ──

def _valid_rrn(seg):
    """주민등록번호: 13자리 + 유효 생년월일(월 01-12, 일 01-31) + 체크섬(가중치 mod-11).
    무작위 13자리 숫자코드를 주민번호로 오인하는 것을 막는다.
    (2020-10 이후 발급분은 뒷자리가 임의번호라 체크섬이 없을 수 있으나,
     '주민번호' 헤더 컬럼은 헤더 규칙으로 마스킹되므로 영향 없음.)"""
    d = re.sub(r"\D", "", seg)
    if len(d) != 13:
        return False
    mm, dd = int(d[2:4]), int(d[4:6])
    if not (1 <= mm <= 12 and 1 <= dd <= 31):
        return False
    weights = (2, 3, 4, 5, 6, 7, 8, 9, 2, 3, 4, 5)
    total = sum(int(d[i]) * weights[i] for i in range(12))
    check = (11 - (total % 11)) % 10
    return check == int(d[12])

def _valid_card(seg):
    """신용카드: 16자리 + Luhn 체크섬 통과."""
    digits = [int(c) for c in re.sub(r"\D", "", seg)]
    if len(digits) != 16:
        return False
    total = 0
    for i, x in enumerate(reversed(digits)):
        if i % 2 == 1:
            x *= 2
            if x > 9:
                x -= 9
        total += x
    return total % 10 == 0

# ── 엔티티 규칙 테이블 (형식 기반 · 정책은 이 표 한 곳에서 관리) ─
# validate(선택): 정규식 매칭 후 추가 검증을 통과해야 인정(오탐 차단).
ENTITY_RULES = [
    {"type": "주민등록번호", "grade": "C", "priority": 1,
     "regex": re.compile(r"\b\d{6}[-\s]?[1-8]\d{6}\b"), "mask": _mask_all_digits,
     "validate": _valid_rrn,
     "why": "주민등록번호 형식 + 생년월일 + 체크섬(검증식) 모두 통과"},
    {"type": "여권번호", "grade": "C", "priority": 2,
     "regex": re.compile(r"\b[MSRODmsrod]\d{8}\b"), "mask": _mask_passport,
     "why": "여권번호 형식(영문 1자+숫자 8자)에 일치"},
    {"type": "신용카드번호", "grade": "C", "priority": 3,
     "regex": re.compile(r"\b\d{4}[-\s]?\d{4}[-\s]?\d{4}[-\s]?\d{4}\b"), "mask": _mask_all_digits,
     "validate": _valid_card,
     "why": "신용카드 형식(4-4-4-4) + Luhn 체크섬 통과"},
    {"type": "운전면허번호", "grade": "C", "priority": 4,
     # 구분자 필수: 맨숫자 12자리(공사/용역번호 등)를 면허로 오인하지 않도록
     "regex": re.compile(r"\b\d{2}[-\s]\d{2}[-\s]\d{6}[-\s]\d{2}\b"), "mask": _mask_all_digits,
     "why": "운전면허 형식(2-2-6-2, 구분자 포함)에 일치"},
    {"type": "계좌번호", "grade": "C", "priority": 5,
     "regex": re.compile(r"(?:계좌(?:번호)?|예금주?|입금|account)\s*[:：\-]?\s*\d[\d\-\s]{7,}\d"),
     "mask": _mask_all_digits, "why": "'계좌/예금/입금' 문맥 + 계좌번호 형식에 일치"},
    {"type": "이메일", "grade": "S", "priority": 6,
     "regex": re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}"), "mask": _mask_email,
     "why": "이메일 형식(local@domain)에 일치"},
    {"type": "전화번호", "grade": "S", "priority": 7,
     # 유효 국번으로 시작할 때만: 휴대폰(01[016789])·서울(02)·지역(0[3-6][1-5])·070·050X
     "regex": re.compile(r"\b0(?:1[016789]|2|[3-6][1-5]|70|50\d)[-\s]?\d{3,4}[-\s]?\d{4}\b"),
     "mask": _mask_phone,
     "why": "전화/휴대폰 형식(유효 국번으로 시작 3-4-4)에 일치"},
    {"type": "주소", "grade": "S", "priority": 8,
     "regex": re.compile(_ADDR_REGION +
                         r"(?:\s*[가-힣A-Za-z0-9]+(?:시|군|구|읍|면|동|리|로|길|가|호|층|번지))+"
                         r"(?:\s*[\d\-]+)?"),
     "mask": _mask_address, "why": "시/도 + 행정구역(구·동·로 등) 패턴에 일치"},
    {"type": "차량번호", "grade": "S", "priority": 9,
     "regex": re.compile(r"\b\d{2,3}[가-힣]\d{4}\b"), "mask": _mask_vehicle,
     "why": "차량번호 형식(숫자+한글+숫자4)에 일치"},
    # 생년월일은 셀 패턴으로 잡지 않는다(접수일·처리일 등 업무 날짜 과마스킹 방지).
    # '생년월일/생일/출생' 제목 칸일 때만 HEADER_RULES로 마스킹한다.
]

# ── 이름 탐지 (설명 가능한 보수적 규칙만 사용) ───────────────────
_SURNAME_2 = ["남궁", "황보", "제갈", "선우", "독고", "서문", "사공", "동방"]
_SURNAME_1 = ("김이박최정강조윤장임한오서신권황안송전홍고문양손배백허유노하"
              "곽성차주우구민진지엄채원천방공현함변염여추도소석선설마길연위표명기반라왕")
_NAME_CORE = r"(?:" + "|".join(_SURNAME_2) + r"|[" + _SURNAME_1 + r"])[가-힣]{1,2}"
_NAME_RE = re.compile(r"(?<![가-힣])" + _NAME_CORE + r"(?![가-힣])")
_NAME_FULL_RE = re.compile(r"^" + _NAME_CORE + r"$")

_HONORIFICS = ("님", "씨", "군", "양", "귀하", "과장", "부장", "차장", "대리", "팀장",
               "사원", "이사", "대표", "선생", "교수", "박사", "주임", "실장", "원장",
               "사장", "회장", "상무", "전무", "고객님", "환자", "학생", "선수", "기사")
_NAME_HEADERS = ("이름", "성명", "성함", "담당자", "담당", "대표자", "신청인", "신청자",
                 "수신자", "수신인", "고객명", "회원명", "환자명", "예금주", "명의자",
                 "가입자", "계약자", "성씨", "name")
_NAME_STOP = {"전화", "전화번호", "연락", "연락처", "주민", "주민번호", "차량", "여권",
              "계좌", "주소", "고객", "회원", "정보", "성명", "이름", "번호", "메일",
              "서울", "부산", "대구", "인천", "광주", "대전", "울산", "세종", "경기",
              "강원", "제주", "영업", "총무", "인사", "재무", "법무", "홍보", "교육",
              "생산", "품질", "구매", "개발", "기획", "구분", "오전", "오후"}
_CONTACT_TYPES = {"전화번호", "이메일", "주민등록번호"}
_GAP_RE = re.compile(r"[\s,/·:()]*")


def _find_names(s, fmt_spans, name_column):
    """설명 가능한 근거가 있을 때만 이름으로 인정한다."""
    names = []
    for mo in _NAME_RE.finditer(s):
        tok, st, en = mo.group(), mo.start(), mo.end()
        if any(not (en <= fs or st >= fe) for fs, fe, *_ in fmt_spans):
            continue                                   # 형식 엔티티와 겹치면 제외
        why = None
        if name_column:
            why = "이름 컬럼으로 판정된 컬럼의 값"
        elif s[en:].lstrip(" ,/·:()").startswith(_HONORIFICS):
            why = "이름 뒤 호칭/직함 동반"
        elif tok not in _NAME_STOP:
            for fs, fe, _g, _p, ftype, _m, _w in fmt_spans:
                if ftype not in _CONTACT_TYPES:
                    continue
                if en <= fs and _GAP_RE.fullmatch(s[en:fs]):
                    why = "연락처/식별정보와 인접"; break
                if fe <= st and _GAP_RE.fullmatch(s[fe:st]):
                    why = "연락처/식별정보와 인접"; break
        if why:
            names.append([st, en, "S", 99, "이름", _mask_name(tok), why])
    return names


def mask_cell(value, name_column=False):
    """
    셀 하나를 스캔해 모든 엔티티를 탐지·마스킹한다.
    반환: (마스킹된 값, [(유형, 등급, 근거), ...])
    탐지가 없으면 원본 값을 그대로 돌려준다(숫자형 등 보존).
    """
    s = str(value)
    fmt = []
    for rule in ENTITY_RULES:
        validate = rule.get("validate")
        for mo in rule["regex"].finditer(s):
            seg = mo.group()
            if validate and not validate(seg):     # 형식은 맞아도 검증 실패 시 제외
                continue
            fmt.append([mo.start(), mo.end(), rule["grade"], rule["priority"],
                        rule["type"], rule["mask"](seg), rule["why"]])
    matches = fmt + _find_names(s, fmt, name_column)
    if not matches:
        return value, []

    matches.sort(key=lambda m: (GRADE_RANK[m[2]], m[3], -(m[1] - m[0])))
    occupied = [False] * len(s)
    chosen = []
    for m in matches:
        st, en = m[0], m[1]
        if any(occupied[st:en]):
            continue
        chosen.append(m)
        for i in range(st, en):
            occupied[i] = True

    chosen.sort(key=lambda m: m[0])
    out, idx, detections = [], 0, []
    for st, en, grade, pr, typ, masked, why in chosen:
        out.append(s[idx:st])
        out.append(masked)
        idx = en
        detections.append((typ, grade, why))
    out.append(s[idx:])
    return "".join(out), detections


def is_name_column(col_name, values):
    """헤더 키워드 또는 값 다수결로 '이름 컬럼' 판정. (values: 셀 값 리스트)"""
    h = str(col_name).lower()
    if any(k.lower() in h for k in _NAME_HEADERS):
        return True
    sample = [str(v).strip() for v in values if v is not None][:200]
    if len(sample) < 5:
        return False
    hit = sum(1 for v in sample if _NAME_FULL_RE.match(v))
    distinct = len(set(sample))
    # 이름 컬럼은 (a) 다수가 성씨+이름 형태이고 (b) 값이 충분히 다양해야 함.
    # (주택/비주택, 전화/방문 같은 '카테고리' 컬럼 오탐 방지)
    return hit >= max(3, len(sample) * 0.6) and distinct >= max(5, len(sample) * 0.3)


def analyze_column(values, name_column, sample_size=200):
    """컬럼 표본 분석 → (대표등급, {유형:건수}, {유형:{근거…}}, 원본예시, 마스킹예시).
    values: 셀 값 리스트."""
    type_counts, type_whys, grade_set = {}, {}, set()
    before = after = ""
    for v in [x for x in values if x is not None][:sample_size]:
        v = str(v)
        masked, dets = mask_cell(v, name_column)
        if dets:
            for typ, g, why in dets:
                type_counts[typ] = type_counts.get(typ, 0) + 1
                type_whys.setdefault(typ, set()).add(why)
                grade_set.add(g)
            if not before:
                before, after = v, str(masked)
    if not type_counts:
        return "O", {}, {}, "", ""
    return ("C" if "C" in grade_set else "S"), type_counts, type_whys, before, after


# ── 헤더(컬럼명) 기반 규칙 — 정형 자료에서 '제목'으로 판별(가장 정확·설명가능) ──
# (헤더에 포함되면 매칭되는 키워드, 유형, 등급, 마스크함수). 위에서부터 우선.
HEADER_RULES = [
    (("주민등록번호", "주민번호"), "주민등록번호", "C", _mask_all_digits),
    (("여권번호",), "여권번호", "C", _mask_passport),
    (("계좌번호", "계좌"), "계좌번호", "C", _mask_all_digits),
    (("카드번호",), "신용카드번호", "C", _mask_all_digits),
    (("운전면허", "면허번호"), "운전면허번호", "C", _mask_all_digits),
    (("사업자등록번호", "사업자번호"), "사업자등록번호", "S", _mask_all_digits),
    (("고객번호",), "고객번호", "S", _mask_all_digits),
    (("우편번호",), "우편번호", "S", _mask_all_digits),
    (("접수번호", "신청번호", "처리번호", "관리번호", "민원번호"), "접수번호", "S", _mask_all_digits),
    (("전화", "휴대폰", "핸드폰", "연락처"), "전화번호", "S", _mask_phone),
    (("이메일", "메일", "email", "e-mail"), "이메일", "S", _mask_email),
    (("주소", "소재지", "거주지"), "주소", "S", _mask_address),
    (("명의자", "신청자", "성명", "이름", "성함", "담당자", "대표자", "수신자", "수신인",
      "고객명", "회원명", "환자명", "예금주", "가입자", "계약자"), "이름", "S", _mask_name),
    (("생년월일", "생일", "출생"), "생년월일", "S", _mask_birth),
    (("차량번호",), "차량번호", "S", _mask_vehicle),
]


def column_masker(header):
    """헤더로 결정되는 (유형, 등급, 마스크함수) 또는 None."""
    h = str(header).replace(" ", "")
    for keys, typ, grade, fn in HEADER_RULES:
        if any(k.replace(" ", "") in h for k in keys):
            return typ, grade, fn
    return None


def grade_of_type(typ):
    for r in ENTITY_RULES:
        if r["type"] == typ:
            return r["grade"]
    for _keys, t, g, _fn in HEADER_RULES:
        if t == typ:
            return g
    return "S"   # 이름 등


# ── 추정 오탐 가능성(%) + 근거 ───────────────────────────────
# 측정된 통계값이 아니라 '규칙 강도'에 기반한 설명 가능한 추정치다.
# 검증 로직이 강할수록 낮고, 휴리스틱일수록 높다. 사람이 근거로 판단할 수 있게 한다.
RISK_HIGH = 30   # 이 값 이상이면 '높음'으로 강조
FP_RISK = {
    "주민등록번호": (2,  "13자리 + 생년월일 + 체크섬(검증식)까지 통과해 오탐이 매우 드뭅니다."),
    "신용카드번호": (3,  "Luhn 체크섬을 통과한 16자리만 인정합니다."),
    "이메일":      (3,  "@ 도메인 형식이 뚜렷합니다."),
    "전화번호":    (7,  "유효 통신 국번(010·02 등)으로 시작하는 번호만 인정합니다."),
    "운전면허번호": (7,  "구분자 포함 2-2-6-2 형식만 인정합니다."),
    "차량번호":    (9,  "숫자+한글+숫자4의 고유 형식입니다."),
    "계좌번호":    (12, "‘계좌/예금/입금’ 문맥이 있을 때만 인정하나 일부 코드와 겹칠 수 있습니다."),
    "주소":       (18, "시/도+행정구역 패턴이라 마스킹 범위가 실제와 다를 수 있습니다."),
    "여권번호":    (24, "영문1+숫자8 형식이 일부 코드와 겹칠 수 있습니다."),
    "생년월일":    (45, "생일이 아닌 일반 날짜(작성일 등)도 같은 형식이라 함께 잡힐 수 있습니다."),
}
# 이름은 '탐지 근거(why)'에 따라 오탐 가능성이 다르므로 별도 표
NAME_RISK = {
    "이름 컬럼으로 판정된 컬럼의 값": (12, "헤더·값 패턴상 이름 컬럼으로 보지만 일부 비(非)인명 값이 섞일 수 있습니다."),
    "이름 뒤 호칭/직함 동반":        (10, "‘님/과장’ 등 직함이 뒤따라 사람일 가능성이 높습니다."),
    "연락처/식별정보와 인접":        (27, "전화·이메일 옆이라 이름으로 추정하나, 성씨로 시작하는 일반 단어일 수 있습니다."),
}


def risk_of(typ, why=None):
    """탐지 1건의 (추정 오탐 %, 근거 문장)."""
    if why and why.startswith("헤더"):
        return (6, "컬럼 제목(헤더)이 해당 정보임을 명시 → 오탐 가능성 낮음.")
    if typ == "이름":
        return NAME_RISK.get(why, (27, "이름으로 추정되나 일반 단어일 수 있습니다."))
    return FP_RISK.get(typ, (20, "자동 판정이라 확인이 필요합니다."))


# ── 데스크톱/웹 공용 고수준 헬퍼 ──────────────────────────────

REPORT_COLUMNS = ["컬럼명", "탐지 정보유형", "N2SF 등급", "건수", "마스킹 정책", "탐지 근거"]


class Table:
    """pandas 없이 엑셀을 다루는 경량 표 구조 (헤더 + 행 리스트)."""
    def __init__(self, headers, rows):
        self.headers = headers      # list[str]
        self.rows = rows            # list[list]

    @property
    def nrows(self):
        return len(self.rows)

    @property
    def ncols(self):
        return len(self.headers)

    def col_index(self, name):
        return self.headers.index(name)   # 동일명 있으면 첫 번째

    def column(self, name):
        i = self.col_index(name)
        return [(r[i] if i < len(r) else None) for r in self.rows]


def _nonempty(row):
    return sum(1 for v in row if v is not None and str(v).strip() not in ("", "-"))


def _build_table(all_rows):
    """전체 행 리스트에서 '헤더 행'을 자동 탐지(제목/빈 줄 건너뜀)하고 Table 생성."""
    all_rows = [r for r in all_rows if r is not None]
    if not all_rows:
        return Table([], [])
    scan = all_rows[:25]
    counts = [_nonempty(r) for r in scan]
    maxc = max(counts) if counts else 0
    if maxc <= 1:
        hdr = 0                                   # 표가 1열뿐이면 그냥 첫 행
    else:
        thresh = max(3, maxc * 0.5)               # 채워진 셀이 충분히 많은 '첫' 행 = 헤더
        hdr = next((i for i, c in enumerate(counts) if c >= thresh), 0)
    headers = [("" if v is None else str(v).strip()) for v in all_rows[hdr]]
    rows = [list(r) for r in all_rows[hdr + 1:]]
    return Table(headers, rows)


def _rows_xlsx(raw):
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    rows = [list(r) for r in wb.active.iter_rows(values_only=True)]
    wb.close()
    return rows


def _rows_xls(raw):
    import xlrd
    book = xlrd.open_workbook(file_contents=raw)
    sh = book.sheet_by_index(0)
    return [list(sh.row_values(r)) for r in range(sh.nrows)]


def _rows_html(raw):
    import pandas as pd
    dfs = pd.read_html(io.BytesIO(raw), header=None)
    if not dfs:
        raise ValueError("HTML에서 표를 찾지 못했습니다.")
    df = dfs[0]
    out = []
    for rec in df.itertuples(index=False, name=None):
        out.append([(None if (v is None or (isinstance(v, float) and v != v)) else v)
                    for v in rec])
    return out


def _from_xlsx(raw):
    return _build_table(_rows_xlsx(raw))


def _from_xls(raw):
    return _build_table(_rows_xls(raw))


def _from_html(raw):
    return _build_table(_rows_html(raw))


def read_table(src):
    """
    파일경로/바이트/파일객체 → Table. 확장자에 의존하지 않고 '실제 내용'으로 판별한다.
    지원: 진짜 xlsx(zip) · 구형 .xls(OLE2) · HTML/XML 표로 위장된 엑셀.
    """
    if isinstance(src, str):
        with open(src, "rb") as f:
            raw = f.read()
    elif isinstance(src, (bytes, bytearray)):
        raw = bytes(src)
    else:
        raw = src.read()
    if not raw:
        return Table([], [])

    low = raw[:2048].lower()
    # 1) 진짜 xlsx (zip 시그니처 PK)
    if raw[:2] == b"PK":
        try:
            return _from_xlsx(raw)
        except Exception:
            pass
    # 2) 구형 .xls (OLE2 시그니처) — 단, 문서보안(DRM) 암호화 파일도 OLE2 컨테이너다
    if raw[:4] == b"\xd0\xcf\x11\xe0":
        try:
            return _from_xls(raw)
        except Exception:
            raise ValueError(
                "구형 .xls가 아니거나 문서보안(DRM)으로 암호화된 파일일 수 있습니다. "
                "Excel에서 파일을 연 뒤 '다른 이름으로 저장 → Excel 통합 문서(*.xlsx)'로 "
                "저장(DRM 해제)한 파일을 사용하세요.")
    # 3) HTML/XML 표 위장
    if (b"<html" in low or b"<table" in low or b"<!doctype" in low
            or low.lstrip().startswith(b"<?xml")):
        return _from_html(raw)
    # 4) 시그니처가 모호하면 순서대로 시도
    for fn in (lambda: _from_xlsx(raw), lambda: _from_xls(raw), lambda: _from_html(raw)):
        try:
            return fn()
        except Exception:
            pass
    raise ValueError(
        "지원하지 않는 형식이거나 손상·암호화된 파일입니다. "
        "문서보안(DRM) 파일이면 Excel에서 열어 '.xlsx'로 다시 저장(해제) 후 사용하세요.")


def _cell(v):
    """openpyxl이 안전하게 기록할 수 있는 값으로 보정."""
    if v is None or isinstance(v, (str, int, float, bool, datetime, date, time)):
        return v
    return str(v)


def _scan_column(header, values, sample_size=200):
    """한 컬럼 분석 → (grade, counts, whys, before, after). 헤더 규칙 우선, 없으면 셀 스캔."""
    rule = column_masker(header)
    name_col = is_name_column(header, values)
    counts, whys, grades = {}, {}, set()
    before = after = ""
    for v in [x for x in values if x is not None][:sample_size]:
        s = str(v)
        if rule:
            typ, grade, fn = rule
            masked = fn(s)
            dets = [(typ, grade, "헤더(컬럼명) 기반 분류")] if masked != s else []
        else:
            masked, dets = mask_cell(s, name_col)
        if dets:
            for t, g, why in dets:
                counts[t] = counts.get(t, 0) + 1
                whys.setdefault(t, set()).add(why)
                grades.add(g)
            if not before:
                before, after = s, str(masked)
    if not counts:
        return "O", {}, {}, "", ""
    return ("C" if "C" in grades else "S"), counts, whys, before, after


def analyze_dataframe(table):
    """전체 표를 컬럼별로 분류. → 리스트[dict]."""
    cols = []
    for col in table.headers:
        grade, counts, whys, before, after = _scan_column(col, table.column(col))
        risk_items = []
        for t in counts:
            pct, reason = max((risk_of(t, w) for w in (whys.get(t) or {None})),
                              key=lambda x: x[0])
            risk_items.append({"type": t, "pct": pct, "reason": reason})
        risk_items.sort(key=lambda x: -x["pct"])
        cols.append({
            "name": str(col), "grade": grade, "types": counts,
            "before": before, "after": after,
            "suggest": grade != "O",
            "risk": max((r["pct"] for r in risk_items), default=0),
            "risk_items": risk_items,
        })
    return cols


def reference_rows():
    """N2SF근거 시트용 (항목, 내용) 행."""
    rows = [("핵심", "본 도구의 핵심은 민감정보 '마스킹'. N2SF는 향후 전환 대비 참고 분류"),
            ("참고 기준", STANDARD_NAME),
            ("배포/시행", STANDARD_TIMELINE + " · 전국 일괄 시행일 미공표(단계적 적용)"),
            ("원문 확인", GUIDE_NOTE),
            ("기밀(C)", f"{GRADE_DEF['C']} → {GRADE_POLICY['C']}"),
            ("민감(S)", f"{GRADE_DEF['S']} → {GRADE_POLICY['S']}"),
            ("공개(O)", f"{GRADE_DEF['O']} → {GRADE_POLICY['O']}")]
    rows += [(f"참고 동향 {i+1}", t) for i, t in enumerate(TREND_NOTES)]
    return rows


def mask_dataframe(table, targets, name_cols=None):
    """
    targets 컬럼을 마스킹. name_cols: {col: bool} (없으면 자동 판정).
    반환: (result_table, report_rows, ref_rows)  ← 시트 기록용 행 리스트(헤더 포함)
    """
    idx_of = {t: table.col_index(t) for t in targets if t in table.headers}
    # 컬럼별 처리 계획: 헤더 규칙('h') 우선, 없으면 셀 스캔('c')
    plan = {}
    for t in idx_of:
        rule = column_masker(t)
        plan[t] = ("h", rule) if rule else ("c", is_name_column(t, table.column(t)))
    counts = {t: {} for t in idx_of}
    whys = {t: {} for t in idx_of}

    new_rows = []
    for row in table.rows:
        nr = list(row)
        for t, i in idx_of.items():
            v = row[i] if i < len(row) else None
            if v is None:
                continue
            s = str(v)
            mode = plan[t]
            if mode[0] == "h":
                typ, _grade, fn = mode[1]
                masked = fn(s)
                if masked != s:
                    nr[i] = masked
                    counts[t][typ] = counts[t].get(typ, 0) + 1
                    whys[t].setdefault(typ, "헤더(컬럼명) 기반 분류")
            else:
                masked, dets = mask_cell(s, mode[1])
                if dets:
                    nr[i] = masked
                    for typ, _g, why in dets:
                        counts[t][typ] = counts[t].get(typ, 0) + 1
                        whys[t].setdefault(typ, why)
        new_rows.append(nr)
    result = Table(table.headers, new_rows)

    report_rows = [list(REPORT_COLUMNS)]
    for t in targets:
        for typ, n in counts.get(t, {}).items():
            g = grade_of_type(typ)
            report_rows.append([t, typ, GRADE_LABEL[g], n, GRADE_POLICY[g], whys[t].get(typ, "")])
    if len(report_rows) == 1:
        report_rows.append(["—", "탐지된 민감정보 없음", "공개(O)", 0, "원본 유지", ""])

    ref_rows = [["항목", "내용"]] + [list(r) for r in reference_rows()]
    return result, report_rows, ref_rows


def count_detections(report_rows):
    """report_rows(헤더 포함)에서 총 마스킹 건수 합계."""
    return sum(int(r[3]) for r in report_rows[1:] if isinstance(r[3], int))


def write_workbook(dst, result_table, report_rows, ref_rows):
    """3개 시트(마스킹결과/분류리포트/N2SF근거)로 dst(경로/파일객체)에 저장."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "마스킹결과"
    ws.append([_cell(h) for h in result_table.headers])
    for r in result_table.rows:
        ws.append([_cell(v) for v in r])
    ws2 = wb.create_sheet("분류리포트")
    for r in report_rows:
        ws2.append([_cell(v) for v in r])
    ws3 = wb.create_sheet("N2SF근거")
    for r in ref_rows:
        ws3.append([_cell(v) for v in r])
    wb.save(dst)
